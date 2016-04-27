## xl.py
## Adam Stein, Patrick Nogacz, Cahle Johnson
## 3-22-2016

from openpyxl import load_workbook
import xlrd

#Test example file
XL_TEST_PATH = 'Docs\Floodlight_Test_1.xlsx'

def xl_parse( file):
	#Parsed Excel File
	parsed_xl = {}

	#Creates a list to later populate with floodlights for output
	lines = []

	#Loads in the workbook file
	xl_book = xlrd.open_workbook( file)
	
	#Selects a workbook sheet
	tsheet = xl_book.sheet_by_name('Traffic Sheet')
	
	#Assign Advertised Id and Profile ID
	profCell = tsheet.cell(8, 1)
	adCell = tsheet.cell(9, 1)
	
	parsed_xl['profileID'] = profCell.value
	parsed_xl['advertiserID'] = adCell.value
	
	## QA for profileID print(parsed_xl['profileID'])
	## QA for advertiserID print(parsed_xl['advertiserID'])
	
	#Loop through Excel File
	for rows in range( 11, tsheet.nrows):
		
		#Reset column to 0
		col = 0
		
		#Reset dictionaries to empty
		line = {}
		
		#Assign values to group lookups
		for cell in tsheet.row(rows):
			col += 1
			
			if(col == 1):
				line['floodlightActivityGroupName'] = cell.value
				
			if(col == 2):
				line['floodlightActivityGroupType'] = cell.value.upper().strip()
			
			if(col == 3):
				line['name'] = cell.value

			if(col == 4):
				line['expectedUrl'] = cell.value
				
			if(col == 5):
				line['standard'] = cell.value.upper().strip()
				
			if(col == 6):
				line['unique'] = cell.value.upper().strip()
				lines.append(line)
				
			
		# QA for groups: print(group)
		# QA for activities: print(activity)
			
	parsed_xl['lines'] = lines
	return parsed_xl

			
	
if __name__ == '__main__':
	xl_parsed = xl_parse( XL_TEST_PATH)
	print(xl_parsed['lines'])