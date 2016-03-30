## xl.py
## Adam Stein, Patrick Nogacz, Cahle Johnson
## 3-22-2016

from openpyxl import load_workbook
import xlrd

#Test example file
XL_TEST_PATH = 'Docs\Floodlight_Example.xlsx'

class floodlight:
		def __init__( self, group, activity):
			self.group = group
			self.activity = activity
			
		def getGroup( self):
			return self.group
			
		def getActivity( self):
			return self.activity
			
		def __repr__(self):
			return str(self.group) + " " + str(self.activity)
			
def xl_parse( file):
	#Parsed Excel File
	parsed_xl = {}

	#Creates a list to later populate with floodlights for output
	floodlights = []

	#Loads in the workbook file
	xl_book = xlrd.open_workbook( file)
	
	#Selects a workbook sheet
	tsheet = xl_book.sheet_by_name('Traffic Sheet')
	
	#Assign Advertised Id and Profile ID
	profCell = tsheet.cell(8, 1)
	adCell = tsheet.cell(9, 1)
	
	parsed_xl['profileID'] = profCell.value
	parsed_xl['advertiserID'] = adCell.value
	
	## QA for profileID print(parsed_xl['profileID'])
	## QA for advertiserID print(parsed_xl['advertiserID'])
	
	#Loop through Excel File
	for rows in range( 11, tsheet.nrows):
		
		#Reset column to 0
		col = 0
		
		#Reset dictionaries to empty
		group = {}
		activity = {}
		
		#Assign values to group lookups
		for cell in tsheet.row(rows):
			col += 1
			
			if(col == 1):
				group['name'] = cell.value
				
			if(col == 2):
				group['type'] = cell.value
			
			if(col == 3):
				activity['name'] = cell.value

			if(col == 4):
				activity['expectedURL'] = cell.value
				
			if(col == 5):
				activity['countingMethod'] = cell.value
				floodlights.append(floodlight(group, activity))
			
		# QA for groups: print(group)
		# QA for activities: print(activity)
			
	parsed_xl['floodlights'] = floodlights
	return parsed_xl

			
	
if __name__ == '__main__':
	floodlights = xl_parse( XL_TEST_PATH)
	print(floodlights['floodlights'])